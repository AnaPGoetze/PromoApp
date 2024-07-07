import openpyxl

class Planilha:
    def __init__(self):
        try:
            self.wb = openpyxl.load_workbook("promocoes.xlsx")
            self.ws = self.wb.active
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Promos"
            self.ws.append(["Local", "Produto", "Desconto", "Latitude", "Longitude"])
            self.wb.save("promocoes.xlsx")

    def inserirPromocaoPlanilha(self, nome_local, nome_produto, desconto, latitude, longitude):
        self.ws.append([nome_local, nome_produto, desconto, latitude, longitude])
        self.wb.save("promocoes.xlsx")


    def lerPlanilha(self):
        return self.ws

    def fecharPlanilha(self):
        self.wb.close()

    #def para remover uma promo, a empresa vai digitar o nome e produto que ele pretende deletar, e toda a linha x da promoção na planilha vai ser deletada
    def removerPromoPlanilha(self, nome_local, nome_produto):
        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == nome_local and row[1].value == nome_produto:
                self.ws.delete_rows(row[0].row)
                self.wb.save("promocoes.xlsx")
                break

    #Def para copiar toda a planilha, e inserir na treeview do tkinter, que está na classTelaExplorar.py
    def copiarPlanilha(self):
        copia = []
        for row in self.ws.iter_rows(min_row=2):
            copia.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value])
        return copia