import openpyxl

from src.utils.path_util import path


class Planilha:
    def __init__(self):
        try:
            self.wb = openpyxl.load_workbook(path.planilha_promocao)
            self.ws = self.wb.active
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Promos"
            self.ws.append(["Local", "Produto", "Desconto", "Latitude", "Longitude"])
            self.wb.save(path)

    def inserir_promocao_planilha(self, nome_local, nome_produto, desconto, latitude, longitude):
        self.ws.append([nome_local, nome_produto, desconto, latitude, longitude])
        self.wb.save(path.planilha_promocao)

    def lerPlanilha(self):
        return self.ws

    def fecharPlanilha(self):
        self.wb.close()

    # def para remover uma promo, a empresa vai digitar o nome e produto que ele pretende deletar, e toda a linha x
    # da promoção na planilha vai ser deletada
    def removerPromoPlanilha(self, nome_local, nome_produto):
        for row in self.ws.iter_rows(min_row=2):
            if row[0].value == nome_local and row[1].value == nome_produto:
                self.ws.delete_rows(row[0].row)
                self.wb.save(path.planilha_promocao)
                break

    # Def para copiar toda a planilha, e inserir na treeview do tkinter, que está na classTelaExplorar.py
    def copiarPlanilha(self):
        copia = []
        for row in self.ws.iter_rows(min_row=2):
            copia.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value])
        return copia

    # Def para procurar uma promo na planilha, o usuario vai digitar o nome do produto,e todas as linhas que tiverem o produto ou parte do produto, as linhas todas vão ser copiadas e inseridas na treeview
    def procurarPromoOupartePlanilha(self, nome_produto):
        copia = []
        for row in self.ws.iter_rows(min_row=2):
            if row[1].value is None:
                continue
            if nome_produto in row[1].value:
                copia.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value])
        return copia
